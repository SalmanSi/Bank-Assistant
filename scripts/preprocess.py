"""
Data Preprocessing module for the NUST Bank Assistant.

Reads the raw Excel knowledge base, cleans messy formatting, unmerges cells,
extracts structured Data (FAQs, rates, limits), detects and logs PII occurrences,
and outputs an array of structured JSON documents ready for embedding.
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

PROJECT_ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = PROJECT_ROOT / "NUST Bank-Product-Knowledge.xlsx"
OUTPUT_PATH = PROJECT_ROOT / "data/processed/documents.json"
PII_SCAN_PATH = PROJECT_ROOT / "data/processed/pii_scan.json"
RATE_SHEET_NAME = "Rate Sheet July 1 2024"
SKIP_SHEETS = {"Main", RATE_SHEET_NAME, "Sheet1"}

CNIC_PATTERN = re.compile(r"\b\d{5}-\d{7}-\d\b")
PHONE_PATTERN = re.compile(r"\b\d{4}-\d{7}\b")
EMAIL_PATTERN = re.compile(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", re.IGNORECASE)
URL_PATTERN = re.compile(r"https?://\S+", re.IGNORECASE)
QUESTION_PATTERN = re.compile(r"^\s*(?:q[:.]\s*|\d+\.\s+).+", re.IGNORECASE)
NUMBERING_PATTERN = re.compile(r"^\s*\d+(?:\.0+)?\s*$")
HEADER_KEYWORDS = {
    "profit payment",
    "profit rate",
    "tenor",
    "payout",
    "minimum age",
    "minimum age to qualify",
    "profit calculation",
    "profit payment frequency",
    "currency",
    "tier 1",
    "tier 2",
    "tier 3",
}

MAIN_TITLE_TO_SHEET = {
    "nust asaan account naa": "NAA",
    "little champs account": "LCA",
    "nust sahar accounts": "NSA",
    "nust waqaar account": "NWA",
    "pakwatan remittance account": "PWRA",
    "nust home remittance": "HOME REMITTANCE",
    "rda digital customer onboarding": "RDA",
    "current deposit account cda": "CDA",
    "value plus current account individual vpca": "VPCA",
    "value plus business account vp ba": "VP-BA",
    "value premium business account vpba": "VPBA",
    "nust maximiser account": "NMA",
    "nust special deposit account nsda": "NSDA",
    "profit and loss sharing account pls": "PLS",
    "nust asaan digital account": "NADA",
    "nust asaan digital remittance account": "NADRA",
    "nust freelancer digital account": "NFDA",
    "nust4car nust bank s auto finance facility faqs": "NUST4Car",
    "personal finance": "PF",
    "nust master card": "NMC",
    "nust mortgage finance": "NMF",
    "nust sahar finance": "NSF",
    "nust imarat finance": "NIF",
    "nust ujala finance": "NUF",
    "nust flour mill finance": "NFMF",
    "nust fauri business finance": "NFBF",
    "prime minister youth business agriculture loan scheme": "PMYB &ALS",
    "nust hunarmand finance": "NHF",
    "nust rice finance": "NRF",
    "nust life bancassurance policy": "Nust Life",
    "efu life bancassurance policy": "EFU Life",
    "jubilee life bancassurance policy": "Jubilee Life ",
}

FALLBACK_SHEET_CATEGORIES = {
    "LCA": "liability",
    "NAA": "liability",
    "NWA": "liability",
    "PWRA": "liability",
    "RDA": "liability",
    "VPCA": "liability",
    "VP-BA": "liability",
    "VPBA": "liability",
    "NSDA": "liability",
    "PLS": "liability",
    "CDA": "liability",
    "NMA": "liability",
    "NADA": "liability",
    "NFDA": "liability",
    "NSA": "liability",
    "NUST4Car": "consumer",
    "PF": "consumer",
    "NMC": "consumer",
    "NMF": "consumer",
    "NSF": "consumer",
    "NIF": "consumer",
    "NUF": "consumer",
    "NFMF": "consumer",
    "NFBF": "consumer",
    "PMYB &ALS": "consumer",
    "NRF": "consumer",
    "NHF": "consumer",
    "ESFCA": "consumer",
    "Nust Life": "insurance",
    "EFU Life": "insurance",
    "Jubilee Life ": "insurance",
    "HOME REMITTANCE": "remittance",
    "NADRA": "remittance",
}


def normalize_key(text: str) -> str:
    """Normalize text into a unified lowercase keyword representation for map matching."""
    lowered = text.lower().replace("&", " ")
    lowered = re.sub(r"[^a-z0-9]+", " ", lowered)
    return re.sub(r"\s+", " ", lowered).strip()


def clean_text(value: Any, *, remove_urls: bool = True) -> str:
    """Sanitize and unify strings (strips whitespaces, transforms bullets)."""
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").replace("\t", " ")
    text = text.replace("•", "- ").replace("·", "- ")
    text = re.sub(r"(?m)^\s*o\s+", "- ", text)
    text = re.sub(r"(?m)^\s*[•·]\s*", "- ", text)
    if remove_urls:
        text = URL_PATTERN.sub("", text)
    text = re.sub(
        r"\bCustomerServices@NUSTbank\.com(?:\.pk)?\b",
        "[BANK_CONTACT_EMAIL]",
        text,
        flags=re.IGNORECASE,
    )
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines()]
    lines = [line for line in lines if line]
    return "\n".join(lines).strip()


def is_navigation_label(text: str) -> bool:
    return normalize_key(text) == "main"


def is_question(text: str) -> bool:
    cleaned = clean_text(text, remove_urls=False)
    return "?" in cleaned or bool(QUESTION_PATTERN.match(cleaned))


def is_numbering(text: str) -> bool:
    return bool(NUMBERING_PATTERN.match(clean_text(text, remove_urls=False)))


def format_rate(value: Any) -> str:
    if isinstance(value, (int, float)):
        numeric = float(value)
        if numeric <= 1:
            numeric *= 100
        return f"{numeric:.2f}%"
    return clean_text(value)


def ensure_parent_dir(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def dedupe_preserve_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    deduped: list[str] = []
    for value in values:
        if value not in seen:
            deduped.append(value)
            seen.add(value)
    return deduped


def load_bank_workbook(workbook_path: Path | str = WORKBOOK_PATH) -> Workbook:
    return load_workbook(workbook_path, data_only=True)


def scan_pii(workbook: Workbook) -> dict[str, list[dict[str, Any]]]:
    findings: dict[str, list[dict[str, Any]]] = {
        "cnic_hits": [],
        "personal_phone_hits": [],
        "personal_email_hits": [],
        "bank_email_hits": [],
        "url_hits": [],
    }
    for ws in workbook.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                raw = str(cell.value)
                location = {"sheet": ws.title, "cell": cell.coordinate, "value": raw}
                if CNIC_PATTERN.search(raw):
                    findings["cnic_hits"].append(location)
                if PHONE_PATTERN.search(raw):
                    findings["personal_phone_hits"].append(location)
                for email in EMAIL_PATTERN.findall(raw):
                    if "nustbank.com" in email.lower():
                        findings["bank_email_hits"].append({**location, "email": email})
                    else:
                        findings["personal_email_hits"].append({**location, "email": email})
                for url in URL_PATTERN.findall(raw):
                    findings["url_hits"].append({**location, "url": url})
    return findings


def write_pii_summary(summary: dict[str, Any], output_path: Path = PII_SCAN_PATH) -> None:
    ensure_parent_dir(output_path)
    output_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")


def unmerge_and_fill_worksheet(worksheet: Worksheet) -> Worksheet:
    ranges = list(worksheet.merged_cells.ranges)
    for merged_range in ranges:
        top_left_value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
        worksheet.unmerge_cells(str(merged_range))
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                worksheet.cell(row, col).value = top_left_value
    return worksheet


def extract_sheet_title(worksheet: Worksheet) -> str:
    for row in worksheet.iter_rows(values_only=True):
        values = [clean_text(value) for value in row if clean_text(value)]
        values = [value for value in values if not is_navigation_label(value)]
        values = dedupe_preserve_order(values)
        if values:
            return values[0]
    return worksheet.title.strip()


def parse_main_sheet(workbook: Workbook) -> tuple[dict[str, str], dict[str, str]]:
    worksheet = workbook["Main"]
    category_map: dict[str, str] = {}
    product_name_map: dict[str, str] = {}
    right_category = "consumer"
    section_headers = {
        "consumer products": "consumer",
        "sme products": "consumer",
        "third party products": "insurance",
    }

    for row in worksheet.iter_rows():
        left_title = clean_text(row[2].value) if len(row) > 2 else ""
        right_header = clean_text(row[4].value) if len(row) > 4 else ""
        right_title = clean_text(row[5].value) if len(row) > 5 else ""

        for candidate in (right_header, right_title):
            normalized = normalize_key(candidate)
            if normalized in section_headers:
                right_category = section_headers[normalized]

        if left_title and not is_numbering(left_title):
            sheet_name = MAIN_TITLE_TO_SHEET.get(normalize_key(left_title))
            if sheet_name:
                category_map[sheet_name] = FALLBACK_SHEET_CATEGORIES.get(sheet_name, "liability")
                product_name_map[sheet_name] = left_title.strip()

        if right_title and not is_numbering(right_title):
            normalized_right = normalize_key(right_title)
            if normalized_right not in section_headers:
                sheet_name = MAIN_TITLE_TO_SHEET.get(normalized_right)
                if sheet_name:
                    category_map[sheet_name] = FALLBACK_SHEET_CATEGORIES.get(sheet_name, right_category)
                    product_name_map[sheet_name] = right_title.strip()

    for sheet_name in workbook.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        category_map.setdefault(sheet_name, FALLBACK_SHEET_CATEGORIES.get(sheet_name, "liability"))
        product_name_map.setdefault(sheet_name, extract_sheet_title(workbook[sheet_name]))

    return category_map, product_name_map


def looks_like_header_row(values: list[str]) -> bool:
    if len(values) < 2 or len(values) > 6:
        return False
    lowered = [normalize_key(value) for value in values]
    if any(value in HEADER_KEYWORDS for value in lowered):
        return True
    return all(not any(char.isdigit() for char in value) for value in values) and all(len(value) <= 35 for value in values)


def format_answer_row(values: list[str]) -> str:
    if not values:
        return ""
    if len(values) == 1:
        return values[0]
    if len(values) == 2:
        left, right = values
        if len(left) <= 50:
            return f"{left}: {right}"
    return "; ".join(values)


def format_tabular_row(headers: list[str], values: list[str]) -> str:
    parts: list[str] = []
    for header, value in zip(headers, values):
        header_text = clean_text(header)
        value_text = clean_text(value)
        if not header_text or not value_text:
            continue
        if "rate" in header_text.lower() and isinstance(value, str):
            value_text = format_rate(value)
        parts.append(f"{header_text}: {value_text}")
    return "; ".join(parts)


def create_document(
    doc_id: str,
    product: str,
    sheet: str,
    doc_type: str,
    category: str,
    content: str,
    question: str | None = None,
) -> dict[str, Any]:
    """Helper struct builder for a document."""
    return {
        "id": doc_id,
        "product": product,
        "sheet": sheet,
        "type": doc_type,
        "category": category,
        "question": question,
        "content": content,
    }


def parse_product_sheet(worksheet: Worksheet, *, product_name: str, category: str) -> list[dict[str, Any]]:
    unmerge_and_fill_worksheet(worksheet)
    documents: list[dict[str, Any]] = []
    document_counter = 0
    title = clean_text(product_name or extract_sheet_title(worksheet), remove_urls=False)
    current_question: str | None = None
    answer_parts: list[str] = []
    general_parts: list[str] = []
    active_headers: list[str] | None = None

    def next_doc_id() -> str:
        nonlocal document_counter
        document_counter += 1
        return f"{worksheet.title}_{document_counter:03d}"

    def flush_question() -> None:
        nonlocal current_question, answer_parts, active_headers
        if not current_question:
            return
        answer = "\n".join(part for part in answer_parts if part).strip()
        if answer:
            documents.append(
                create_document(
                    next_doc_id(),
                    title,
                    worksheet.title,
                    "qa_pair",
                    category,
                    f"Q: {current_question}\nA: {answer}",
                    current_question,
                )
            )
        current_question = None
        answer_parts = []
        active_headers = None

    for row in worksheet.iter_rows(values_only=True):
        cleaned_values = [clean_text(value) for value in row if clean_text(value)]
        cleaned_values = [value for value in cleaned_values if not is_navigation_label(value)]
        cleaned_values = dedupe_preserve_order(cleaned_values)
        if not cleaned_values:
            active_headers = None
            continue

        joined = " ".join(cleaned_values).strip()
        if not joined or joined == title:
            continue

        if is_question(joined):
            flush_question()
            current_question = joined
            continue

        if current_question:
            if looks_like_header_row(cleaned_values):
                active_headers = cleaned_values
                continue
            if active_headers and len(cleaned_values) >= 2:
                answer_parts.append(format_tabular_row(active_headers, cleaned_values))
                continue
            active_headers = None
            answer_parts.append(format_answer_row(cleaned_values))
        else:
            general_parts.append(format_answer_row(cleaned_values))

    flush_question()

    if general_parts:
        content = "\n".join(part for part in general_parts if part).strip()
        if content:
            documents.insert(
                0,
                create_document(
                    next_doc_id(),
                    title,
                    worksheet.title,
                    "general_info",
                    category,
                    content,
                ),
            )
    return documents


def parse_rate_sheet(workbook: Workbook) -> list[dict[str, Any]]:
    worksheet = workbook[RATE_SHEET_NAME]
    documents: list[dict[str, Any]] = []
    counter = 0
    savings_product: str | None = None
    term_product: str | None = None

    def add_rate_document(content: str) -> None:
        nonlocal counter
        counter += 1
        documents.append(
            create_document(
                f"RATE_{counter:03d}",
                "Rate Sheet",
                RATE_SHEET_NAME,
                "rate_info",
                "rate",
                content,
            )
        )

    for row_index in range(1, worksheet.max_row + 1):
        left_name = clean_text(worksheet.cell(row_index, 2).value)
        left_rate = worksheet.cell(row_index, 4).value
        term_name_candidate = clean_text(worksheet.cell(row_index, 6).value)
        payout = clean_text(worksheet.cell(row_index, 7).value)
        term_rate = worksheet.cell(row_index, 9).value

        if left_name and left_name not in {"Profit Payment", "Tenor", "Change", "SAVINGS ACCOUNTS", "TERM DEPOSITS"}:
            if not is_numbering(left_name) and not payout and left_rate in (None, ""):
                savings_product = left_name

        if term_name_candidate and term_name_candidate not in {"Tenor", "Term Deposits", "Profit Rate", "Profit Rate ", "Payout"}:
            if not payout and term_rate in (None, ""):
                term_product = term_name_candidate

        if left_name and left_rate not in (None, "") and left_name != "Profit Payment":
            product = savings_product or "Savings Account"
            add_rate_document(
                f"{product}: Profit payment {left_name} at {format_rate(left_rate)} per annum."
            )

        if term_name_candidate and term_rate not in (None, "") and term_name_candidate != "Tenor":
            product = term_product
            if product in {None, "Term Deposits"} and savings_product:
                product = f"{savings_product} Term Deposit"
            if not product:
                product = "Term Deposit"
            payout_text = f", paid {payout}" if payout else ""
            add_rate_document(
                f"{product}: {term_name_candidate} tenor{payout_text}, rate {format_rate(term_rate)}."
            )

    fcy_headers = [clean_text(worksheet.cell(59, column).value) for column in range(2, 6)]
    if fcy_headers[:4] == ["FCY", "USD", "GBP", "EUR"]:
        for row_index in (60, 61):
            account_type = clean_text(worksheet.cell(row_index, 2).value)
            if not account_type:
                continue
            rates = []
            for offset, currency in enumerate(("USD", "GBP", "EUR"), start=3):
                value = worksheet.cell(row_index, offset).value
                if value not in (None, ""):
                    rates.append(f"{currency} {format_rate(value)}")
            if rates:
                add_rate_document(f"Foreign currency {account_type}: {'; '.join(rates)} per annum.")

    return documents


def preprocess_workbook(
    workbook_path: Path | str = WORKBOOK_PATH,
    output_path: Path | str = OUTPUT_PATH,
    pii_output_path: Path | str = PII_SCAN_PATH,
) -> list[dict[str, Any]]:
    """Master preprocessing pipeline matching all extraction logic and writing the artifacts."""
    workbook = load_bank_workbook(workbook_path)
    pii_summary = scan_pii(workbook)
    write_pii_summary(pii_summary, Path(pii_output_path))

    category_map, product_name_map = parse_main_sheet(workbook)
    documents: list[dict[str, Any]] = []
    documents.extend(parse_rate_sheet(workbook))

    for sheet_name in workbook.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        worksheet = workbook[sheet_name]
        documents.extend(
            parse_product_sheet(
                worksheet,
                product_name=product_name_map.get(sheet_name, extract_sheet_title(worksheet)),
                category=category_map[sheet_name],
            )
        )

    cleaned_documents = [document for document in documents if clean_text(document["content"])]
    output_file = Path(output_path)
    ensure_parent_dir(output_file)
    output_file.write_text(json.dumps(cleaned_documents, indent=2, ensure_ascii=False), encoding="utf-8")
    return cleaned_documents


def main() -> None:
    documents = preprocess_workbook()
    workbook = load_bank_workbook()
    pii_summary = scan_pii(workbook)
    qa_count = sum(1 for document in documents if document["type"] == "qa_pair")
    rate_count = sum(1 for document in documents if document["type"] == "rate_info")
    print(f"Processed {len(documents)} documents")
    print(f"Q&A pairs: {qa_count}")
    print(f"Rate documents: {rate_count}")
    print(
        "PII scan: "
        f"CNIC={len(pii_summary['cnic_hits'])}, "
        f"phones={len(pii_summary['personal_phone_hits'])}, "
        f"personal_emails={len(pii_summary['personal_email_hits'])}"
    )
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()